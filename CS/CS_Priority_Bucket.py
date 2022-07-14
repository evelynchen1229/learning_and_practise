from enum import unique
import pandas as pd
import numpy as np
import os
import sys
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell

#cs_data = sys.path[0] + '\\CS PROPOSED DATA New Business.xlsx'
cs_data = sys.path[0] + '\\CS PN DATA.xlsx'
high_growth = sys.path[0] + '\\Top 10s AM Field 2022.xlsx'
biz_priority = sys.path[0] + '\\Business Priority Accounts.xlsx'

df_cs = pd.read_excel(cs_data)
df_high_growth = pd.read_excel(high_growth,'Combined')
df_biz_priority = pd.read_excel(biz_priority)

#high_growth_account = df_high_growth['Account Number'].to_list()
def spending(x):
    if x['ALL BILLING ACCOUNT ON SPEND'] > 50000:
        return 'High'
    elif 20000 <= x['ALL BILLING ACCOUNT ON SPEND'] <= 50000:
        return  'Medium'
    else: 
        return 'Low'

def cusspending(x):
    if x['ADJUSTED PERIOD END CUSTOMER ON SPEND'] > 50000:
        return 'High'
    elif 20000 <= x['ADJUSTED PERIOD END CUSTOMER ON SPEND'] <= 50000:
        return  'Medium'
    else: 
        return 'Low'



df_priority = (df_cs
.merge(df_high_growth[['Account Number','Upsell Potential (GBP)']],how='left',left_on='ACCOUNT',right_on='Account Number')
.merge(df_biz_priority[['ACCOUNT','BP_>£5k','BP_£3-5k','BP_IH','BP_OTHER','BP_ACCT']],how='left',on='ACCOUNT')
.assign(acct_risk_driver = lambda x: np.where( x['RISK_LVL'].isin(['High','Low']),'High',''))
.assign(acct_high_growth_driver = lambda x: np.where(pd.isna(x['Upsell Potential (GBP)']),'','High') )
.assign(acct_strategic_driver= lambda x: np.where(x['TEAM'].str.contains('ST-'),'High',''))
.assign(acct_new_biz_driver = lambda x: np.where(x['NEW_BIZ_DEALS']==('New Business'),'High',''))
.assign(acct_spending_driver=lambda x: x.apply(lambda x: spending(x),axis=1))
.assign(acct_pn_driver= lambda x: np.where(x['PREMIUM NEWS SPEND FLAG']=='Y','High',''))
.assign(accnt_driver = lambda x: x['BP_ACCT'].fillna('')+x['acct_risk_driver'] + x['acct_high_growth_driver'] + x['acct_strategic_driver'] + x['acct_new_biz_driver'] + x['acct_spending_driver'] + x['acct_pn_driver'])
.assign(accnt_priority = lambda x: np.where(x['accnt_driver'].str.contains('High'),'High',np.where(x['accnt_driver'].str.contains('Medium'),'Medium','Low')))
)


df_cus_risk = (df_priority[['CUSTOMER','acct_risk_driver']].drop_duplicates()
.groupby (['CUSTOMER'])['acct_risk_driver'].apply(', '.join).reset_index()
.rename(columns = {'acct_risk_driver':'acct_risk_driver_combined'})
.assign(cus_risk_driver = lambda x: np.where(x['acct_risk_driver_combined'].str.contains('High'),'High',''))
)

df_cus_hg = (df_priority[['CUSTOMER','acct_high_growth_driver']].drop_duplicates()
.groupby (['CUSTOMER'])['acct_high_growth_driver'].apply(', '.join).reset_index()
.rename(columns = {'acct_high_growth_driver':'acct_high_growth_driver_combined'})
.assign(cus_high_growth_driver = lambda x: np.where(x['acct_high_growth_driver_combined'].str.contains('High'),'High',''))
)
df_cus_strategic=(df_priority[['CUSTOMER','acct_strategic_driver']].drop_duplicates()
.groupby (['CUSTOMER'])['acct_strategic_driver'].apply(', '.join).reset_index()
.rename(columns = {'acct_strategic_driver':'acct_strategic_driver_combined'})
.assign(cus_strategic_driver = lambda x: np.where(x['acct_strategic_driver_combined'].str.contains('High'),'High',''))
)
df_cus_nb=(df_priority[['CUSTOMER','acct_new_biz_driver']].drop_duplicates()
.groupby (['CUSTOMER'])['acct_new_biz_driver'].apply(', '.join).reset_index()
.rename(columns = {'acct_new_biz_driver':'acct_new_biz_driver_combined'})
.assign(cus_new_biz_driver = lambda x: np.where(x['acct_new_biz_driver_combined'].str.contains('High'),'High',''))
)
df_cus_pn=(df_priority[['CUSTOMER','acct_pn_driver']].drop_duplicates()
.groupby (['CUSTOMER'])['acct_pn_driver'].apply(', '.join).reset_index()
.rename(columns = {'acct_pn_driver':'acct_pn_driver_combined'})
.assign(cus_pn_driver = lambda x: np.where(x['acct_pn_driver_combined'].str.contains('High'),'High',''))
)

df_cus_spend = (df_priority[['CUSTOMER','ADJUSTED PERIOD END CUSTOMER ON SPEND']].drop_duplicates()
.assign(cus_spending_driver = lambda x: x.apply(lambda x: cusspending(x),axis=1))
)

df_cus_bp_over_5k = (df_priority[['CUSTOMER','BP_>£5k']].drop_duplicates()
.assign(bp_over_5k = lambda x: x['BP_>£5k'].fillna(''))
.drop(columns = ['BP_>£5k'])
.groupby (['CUSTOMER'])['bp_over_5k'].apply(', '.join)
.reset_index()
.rename(columns = {'bp_over_5k':'bp_over_5k_combined'})
.assign(cus_bp_over_5k = lambda x: np.where(x['bp_over_5k_combined'].str.contains('High'),'High',''))
)


df_cus_bp_3_5k = (df_priority[['CUSTOMER','BP_£3-5k']].drop_duplicates()
.assign(bp_3_5k = lambda x: x['BP_£3-5k'].fillna(''))
.drop(columns = ['BP_£3-5k'])
.groupby (['CUSTOMER'])['bp_3_5k'].apply(', '.join)
.reset_index()
.rename(columns = {'bp_3_5k':'bp_3_5k_combined'})
.assign(cus_bp_3_5k = lambda x: np.where(x['bp_3_5k_combined'].str.contains('High'),'High',''))
)


df_cus_bp_ih = (df_priority[['CUSTOMER','BP_IH']].drop_duplicates()
.assign(bp_ih = lambda x: x['BP_IH'].fillna(''))
.drop(columns = ['BP_IH'])
.groupby (['CUSTOMER'])['bp_ih'].apply(', '.join)
.reset_index()
.rename(columns = {'bp_ih':'bp_ih_combined'})
.assign(cus_bp_ih = lambda x: np.where(x['bp_ih_combined'].str.contains('High'),'High',''))
)


df_cus_bp_other = (df_priority[['CUSTOMER','BP_OTHER']].drop_duplicates()
.assign(bp_other = lambda x: x['BP_OTHER'].fillna(''))
.drop(columns = ['BP_OTHER'])
.groupby (['CUSTOMER'])['bp_other'].apply(', '.join)
.reset_index()
.rename(columns = {'bp_other':'bp_other_combined'})
.assign(cus_bp_other = lambda x: np.where(x['bp_other_combined'].str.contains('High'),'High',''))
)


df_cus_biz = (df_priority[['CUSTOMER','BP_ACCT']].drop_duplicates()
.assign(bp_acct = lambda x: x['BP_ACCT'].fillna(''))
.drop(columns = ['BP_ACCT'])
.groupby (['CUSTOMER'])['bp_acct'].apply(', '.join)
.reset_index()
.rename(columns = {'bp_acct':'bp_acct_combined'})
.assign(cus_bp = lambda x: np.where(x['bp_acct_combined'].str.contains('High'),'High',''))
)
# df_cus_priority = (df_priority[['CUSTOMER','accnt_priority']].drop_duplicates()
# .groupby (['CUSTOMER'])['accnt_priority'].apply(', '.join).reset_index()
# .rename(columns = {'accnt_priority':'accnt_priority_combined'})
# .assign(cus_priority = lambda x: np.where(x['accnt_priority_combined'].str.contains('High'),'High',np.where(x['accnt_priority_combined'].str.contains('Medium'),'Medium','Low')))
# )

# check whether the cus record is unique, no duplication
#check = (df_cus_priority.groupby(['CUSTOMER'])[['accnt_priority_combined','cus_priority']].count().reset_index())[['accnt_priority_combined','cus_priority']].drop_duplicates()
#print(check)
#print(df_cus_priority)

df_priority_final = (df_priority
                 .merge(df_cus_risk,how='left',on='CUSTOMER')
                 .merge(df_cus_hg,how='left',on='CUSTOMER')
                 .merge(df_cus_strategic,how='left',on='CUSTOMER')
                 .merge(df_cus_nb,how='left',on='CUSTOMER')
                 .merge(df_cus_spend[['CUSTOMER','cus_spending_driver']],how='left',on='CUSTOMER')
                 .merge(df_cus_pn,how='left',on='CUSTOMER')
                 .merge(df_cus_bp_over_5k,how='left',on='CUSTOMER')
                 .merge(df_cus_bp_3_5k,how='left',on='CUSTOMER')
                 .merge(df_cus_bp_ih,how='left',on='CUSTOMER')
                 .merge(df_cus_bp_other,how='left',on='CUSTOMER')
                 .merge(df_cus_biz,how='left',on='CUSTOMER')
                 .assign(cus_driver = lambda x: x['cus_bp'] +x['cus_risk_driver'] + x['cus_high_growth_driver'] + x['cus_strategic_driver'] + x['cus_new_biz_driver'] + x['cus_spending_driver'] + x['cus_pn_driver'])
                 .assign(cus_priority = lambda x: np.where(x['cus_driver'].str.contains('High'),'High',np.where(x['cus_driver'].str.contains('Medium'),'Medium','Low')))
                # .merge(df_cus_priority,how='left',on='CUSTOMER')
                 
).drop(columns = ['bp_over_5k_combined','bp_3_5k_combined','bp_ih_combined','bp_other_combined','Account Number','CSM','accnt_driver','cus_driver','acct_risk_driver_combined','acct_high_growth_driver_combined','acct_strategic_driver_combined','acct_new_biz_driver_combined','acct_pn_driver_combined','bp_acct_combined'])
#print(df_priority_final)

# account base by team by priority bucket
priority_base = df_priority_final.groupby(['CSTEAM','cus_priority'])['CUSTOMER'].nunique().unstack().reset_index()
priority_base.columns = priority_base.columns.str.strip().str.upper()
df_priority_final.columns = df_priority_final.columns.str.strip().str.upper()
#print(accnt_base_priority_bucket)



# save report and summary
priority_file = sys.path[0] + '\\CS PRIORITY DATA.xlsx'
sheets = ['Current Account Base','Report']
dfs = [priority_base,df_priority_final]
wb = openpyxl.load_workbook(priority_file)

for df, sheet in zip(dfs, sheets):
    if sheet in wb.sheetnames:
        del wb[sheet]
        wb.save(priority_file)
    writer = pd.ExcelWriter(priority_file, engine='openpyxl')
    writer.book = wb    
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.save()
if 'Sheet1' in wb.sheetnames:
    del wb['Sheet1']
    wb.save(priority_file)

