import pandas as pd
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from datetime import date
import os
import sys

today = date.today().strftime('%Y%m%d')

tax_cols = ['CSM_EMAIL','CSM_NAME','REP_EMAIL','SALES_TERRITORY','REP_NAME','REP_CODE','TOLLEY_DOCVIEWS','TOLLEY_PAGEVIEWS','TOLLEY_LOGINS','GUIDANCE_PAGEVIEWS','GUIDANCE_DOCVIEWS','GUIDANCE_LOGINS','EXC_EMAIL','USERNAME','TOLLEY_MIN_RENEWALDATE','TOLLEY_MIN_STARTDATE','TOLLEY_LIVE','GUIDANCE_RENEWAL_DATE','GUIDANCE_MIN_START_DATE','GUIDANCE_LIVE','GENESIS_ACCOUNT','MASTERNAME','MASTERNUMBER','PRODUCT_NAME','PRODUCT','PRODUCTEMAIL','BUSINESS','MARKET','VALUE','SUB_SEGMENT','SEGMENT_UK','LANGUAGE','REGION','COUNTRY','CONTACT_PHONE','TITLE','USER_LN','USER_FN','SALUTATIONS','GREETING','EMAIL','SITE_ID','ORGANISATION_ID','CONTACT_ID']

legal_cols = ['CSM_EMAIL','CSM_NAME','REP_EMAIL','SALES_TERRITORY','REP_NAME','REP_CODE','LIBRARY_DOCVIEWS','LIBRARY_PAGEVIEWS','LIBRARY_LOGINS','PSL_DOC_VIEWS','PSL_PAGE_VIEWS','PSL_LOGINS','EXC_EMAIL','USERNAME','LIB_MIN_RENEWALDATE','LIB_MIN_STARTDATE','LIBRARY_LIVE','PSL_MIN_RENEWALDATE','PSL_MIN_STARTDATE','PSL_LIVE','GENESIS_ACCOUNT','MASTERNAME','MASTERNUMBER','PRODUCT_NAME','PRODUCT','PRODUCTEMAIL','BUSINESS','MARKET','VALUE','SUB_SEGMENT','CUSTOMER_GROUP','LANGUAGE','REGION','COUNTRY','CONTACT_PHONE','TITLE','USER_LN','USER_FN','SALUTATIONS','GREETING','EMAIL','SITE_ID','ORGANISATION_ID','CONTACT_ID']

output_final = sys.path[0]+'\\3. Raw Data\\final_data_set.xlsx'
df_all = pd.read_excel(output_final)

funnel_file = sys.path[0]+'\\3. Raw Data\\funnel_view_and_summary.xlsx'
funnel_view = pd.read_excel(funnel_file,'Funnel_View')

df_tax = (df_all.loc[lambda x: x['PRODUCT_NAME'].str.contains('Tolley')]
    .rename(columns = {'product':'PRODUCTEMAIL'})
    )[tax_cols]
# convert dates to American format
for col in ['TOLLEY_MIN_RENEWALDATE','TOLLEY_MIN_STARTDATE','GUIDANCE_RENEWAL_DATE','GUIDANCE_MIN_START_DATE']:
    df_tax[col] = df_tax[col].dt.strftime('%m/%d/%Y')

df_legal = (df_all.loc[lambda x: ~x['PRODUCT_NAME'].str.contains('Tolley')]
    .rename(columns = {'product':'PRODUCTEMAIL'})
    )[legal_cols]
for col in ['LIB_MIN_RENEWALDATE','LIB_MIN_STARTDATE','PSL_MIN_RENEWALDATE','PSL_MIN_STARTDATE']:
    df_legal[col] = df_legal[col].dt.strftime('%m/%d/%Y')
print('generating NPS_Main_Data_Tax ...')
output_final_tax = sys.path[0]+r'\\NPS_Main_Data_Tax_{}.xlsx'.format(today)
df_tax.to_excel(output_final_tax,index=False)

print('generating NPS_Main_Data_Legal ...')
output_final_legal = sys.path[0]+r'\\NPS_Main_Data_Legal_{}.xlsx'.format(today)
df_legal.to_excel(output_final_legal,index=False)

print('generating final funnel view and summary file ...')

funnel_view_trans = funnel_view.transpose()
wb = openpyxl.load_workbook(funnel_file)
writer = pd.ExcelWriter(funnel_file, engine='openpyxl')
writer.book = wb
if 'Funnel_View' in wb.sheetnames:
    del wb['Funnel_View']
    wb.save(funnel_file)
funnel_view_trans.to_excel(writer,sheet_name = 'Funnel_View', header=None)
writer.save()

print('NPS Done!')
# formatting dates in American format