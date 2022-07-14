# load modules
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
import os
import sys

# get engagemnet account
engagement_file = sys.path[0]+'\\engagement.txt'
with open(engagement_file) as f:
        lines = f.readlines()
        accnt_engagement = [line.strip() for line in lines]
        f.close()

# load constants
PRODUCTFAMILY = ['LexisLibrary','LexisPSL','TolleyGuidance','TolleyLibrary']
# requested by Sarah Wilson
ACCNT_EXCL = ['EMMS5003','KNIG5010','KNIG5103','KNIG5104']
# requested by Matt Gowmell Jan 2022
# PRIC5402 - Pricewaterhouse Coopers - Ireland
# PRIC5129 - PricewaterhouseCoopers Services Ltd
PWC_EXCL=['PRIC5129','PRIC5402']
ACCNT_INTERNAL =['LEXI5081','WILL5016','WEST5276','WBIR5000','WALK5098','VALE5035','UKGL5000','TRAN5024','THEC6838','THEC5067','TENO5005','SURR5028','STHC5001','SSTA5006','SSTA5005','SOUT6633','SIGS5000','RWGO5000','ROYA5211','RMJA5001','RGKA5001','REED5133','REED0020','PUBL5023','PTEM5000','PSLT5000','PSLS5000','PRIS5017','POLI5001','PHUN5000','PENN5015','PAUL5020','PARR5019','OFFI5316','OFFI5309','NTHW5002','NOTT5048','NDAL5000','MRRI5005','MROD5001','MINI5354','MINI5341','MINI5276','MHAY5002','MATT5103','MANA5042','LUTO5027','LORD5061','LOND6472','LNTE5000','LNBT5001','LNBT5000','LEXI5396','LEXI5390','LEXI5386','LEXI5377','LEXI5365','LEXI5352','LEXI5335','LEXI5329','LEXI5318','LEXI5286','LEXI5278','LEXI5275','LEXI5264','LEXI5239','LEXI5202','LEXI5161','LEXI5126','LEXI5107','LEXI5092','LEXI5008','LEIC5040','LEGA6495','LDJH5000','LANC5039','KNOW5038','KING5475','JUST5047','JUST5022','JUDI5046','JUDG5086','JUDG5084','JUDG5082','JUDG5081','JUDG5080','JUDG5077','JUDG5024','JUDG5011','JUDG0021','JSCA5001','JOHN5599','JOHN5584','JMIL5007','JHRE5004','JDRA5000','JAME5753','JAME5750','INTE5697','HOWA5063','HONM5083','HONM5069','HONM5058','HONM5046','HONM5038','HONM5037','HONM5012','HONJ5005','HMPA5000','HMCO5003','HISH5132','HISH5131','HISH5129','HISH5127','HISH5126','HISH5125','HISH5122','HISH5121','HISH5120','HISH5118','HISH5110','HISH5109','HISH5108','HISH5107','HISH5106','HISH5105','HISH5102','HISH5101','HISH5099','HISH5097','HISH5094','HISH5092','HISH5087','HISH5084','HISH5083','HISH5080','HISH5078','HISH5076','HISH5067','HISH5055','HISH5054','HISH5046','HISH5044','HISH5042','HISH5030','HISH5029','HISH5028','HISH5027','HISH5024','HISH5022','HISH5020','HISH5015','HISH5009','HISH5007','HHJU5000','HERH5006','HERH5004','HERH5001','HENR5091','HCLT5002','HAMM0028','HAM19945','HAM19153','GRAT7646','GRAT5001','GOSN5000','GMCG0001','GLYN5000','GLPT5000','EXTE5005','EDWA5157','EBEN5001','DORS5025','DIST5060','DIST5045','DIST5036','DIST5020','DIST5016','DIST5011','DIST5007','DIST5006','DIST5005','COUR6017','COMP5248','CLIF5132','CLER0069','CHRI5524','CHRI5098','CHAI0013','CBLO5000','CARO5151','BUTT5073','BUTT5020','BLEV5003','BJLA5000','BEAS5004','BAKE5016','AUTH5003','AUDI5004','ATHE5020','ASIM5004','ARMY5003','APGH5001','AMBU5000','AJSE5000','5THB0001','LNIN5001','TREV5076']
NAME_INTERNAL = ['GRAT','AUTH','STUD','LEXI','LEGA','BUTT','EXTE','HAMM']
EMAIL_INTERNAL = ['clerk','student','test','trainee','training','workexp','library','public','lawsociety']
EMAIL_SENSITIVE = ['ministryofjustice','lexisnex','xperth','justice.gov.uk']
#requested by David Smith
#exclude the email for fiona as she might be flagged in the wrong account
EMAIL_EXCL = ['francesca.hutcheson@dykeyaxley.co.uk','fiona.wallace@leiperandsummers.co.uk']



# get dataframes ready
# for tax customers
output_tax_usg =  sys.path[0]+'\\3. Raw Data\\NPS-Tax_Usage.xlsx'
tax_usg = pd.read_excel(output_tax_usg)

output_tax_nps = sys.path[0]+'\\3. Raw Data\\NPS-Tax.xlsx'
tax_nps = pd.read_excel(output_tax_nps)

output_tax_email = sys.path[0]+'\\3. Raw Data\\tax_hub_no_email.xlsx'
tax_nps_null_email = pd.read_excel(output_tax_email)

# for legal customers
output_legal_usg =  sys.path[0]+'\\3. Raw Data\\NPS-Legal_Usage.xlsx'
legal_usg = pd.read_excel(output_legal_usg)

output_legal_nps = sys.path[0]+'\\3. Raw Data\\NPS-Legal.xlsx'
legal_nps = pd.read_excel(output_legal_nps)

output_legal_email = sys.path[0]+'\\3. Raw Data\\legal_hub_no_email.xlsx'
legal_nps_null_email = pd.read_excel(output_legal_email)

# nps funnel view file
step1_pipeline = (pd.concat([tax_usg,legal_usg])
                  [['GENESIS_ACCOUNT','MASTERNAME','USERNAME','EMAIL','PRODUCT_NAME']]
                  .drop_duplicates()
                  .groupby('PRODUCT_NAME')['USERNAME'].nunique()
                  .reset_index()
                  .rename(columns = {'PRODUCT_NAME':'product','USERNAME':'contacts_step1'} )
                 )
step2_pipeline = (pd.concat([tax_nps,legal_nps])
                  .groupby('PRODUCTEMAIL')['USERNAME'].nunique()
                  .reset_index()
                  .rename(columns = {'PRODUCTEMAIL':'product','USERNAME':'contacts_step2'} )
                 )
drop1_pipeline = (pd.concat([tax_nps,legal_nps,tax_nps_null_email ,legal_nps_null_email])
                  .groupby('PRODUCTEMAIL')['USERNAME'].nunique()
                  .reset_index()                  
                  .rename(columns={'PRODUCTEMAIL':'product','USERNAME':'contacts_drop1'}) 
                 )
# first dropout : users who satisfy usage threshold are not in the customer hub
hub_dropout = (step1_pipeline
               .merge(drop1_pipeline,how='inner',on='product',validate = 'one_to_one')
               .assign(hub_dropout = lambda x:x['contacts_step1'] - x['contacts_drop1'])
              )
# second dropout: users who are in customer hub don't have email address
email_dropout = (hub_dropout
                 .merge(step2_pipeline,how='left',on='product',validate='one_to_one')
                 .assign(hub_email_dropout = lambda x: x['contacts_drop1'] - x['contacts_step2'])
                )
# third dropout: accounts requested by Sarah Wilson should be excluded
accnt_excl_drop = (pd.concat([tax_nps,legal_nps])
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(ACCNT_EXCL)]
                   .groupby('PRODUCTEMAIL')['USERNAME'].nunique()
                   .reset_index()
                   .rename(columns={'PRODUCTEMAIL':'product','USERNAME':'contacts_drop2'})
                  )
accnt_dropout = (email_dropout
                 .merge(accnt_excl_drop,how='left',on='product',validate='one_to_one')
                 .assign(sarah_accnt_dropout = lambda x: x['contacts_step2'] - x['contacts_drop2'])
                )
# fourth dropout: engagement accounts should be excluded
engagement_drop = (pd.concat([tax_nps,legal_nps])
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(ACCNT_EXCL)]
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(accnt_engagement)]
                   .groupby('PRODUCTEMAIL')['USERNAME'].nunique()
                   .reset_index()
                   .rename(columns={'PRODUCTEMAIL':'product','USERNAME':'contacts_drop3'})
                  )
engagement_dropout = (accnt_dropout
                 .merge(engagement_drop,how='left',on='product',validate='one_to_one')
                 .assign(engagement_dropout = lambda x: x['contacts_drop2'] - x['contacts_drop3'])
                     )
# fifth dropout: internal and sensitive accounts and users should be excluded
internal_drop = (pd.concat([tax_nps,legal_nps])
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(ACCNT_EXCL)]
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(accnt_engagement)]
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(ACCNT_INTERNAL)]
                   .loc[lambda x: ~x['EMAIL'].str.contains('|'.join(EMAIL_SENSITIVE))]
                   .loc[lambda x: ~(x['EMAIL'].str.contains('|'.join(EMAIL_INTERNAL)) & x['USERNAME'].str.contains('@'))]
                   .groupby('PRODUCTEMAIL')['USERNAME'].nunique()
                   .reset_index()
                   .rename(columns={'PRODUCTEMAIL':'product','USERNAME':'contacts_drop4'})
                  )
internal_dropout = (engagement_dropout
                 .merge(internal_drop,how='left',on='product',validate='one_to_one')
                 .assign(internal_dropout = lambda x: x['contacts_drop3'] - x['contacts_drop4'])
                   )
# sixth dropout: emails requested by David Smith should be excluded
email_excl_drop = (pd.concat([tax_nps,legal_nps])
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(ACCNT_EXCL)]
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(accnt_engagement)]
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(ACCNT_INTERNAL)]
                   .loc[lambda x: ~x['EMAIL'].str.contains('|'.join(EMAIL_SENSITIVE))]
                   .loc[lambda x: ~(x['EMAIL'].str.contains('|'.join(EMAIL_INTERNAL)) & x['USERNAME'].str.contains('@'))]
                   .loc[lambda x: ~x['EMAIL'].isin(EMAIL_EXCL)]
                   .groupby('PRODUCTEMAIL')['USERNAME'].nunique()
                   .reset_index()
                   .rename(columns={'PRODUCTEMAIL':'product','USERNAME':'contacts_drop5'})
                  )
email_excl_dropout = (internal_dropout
                 .merge(email_excl_drop,how='left',on='product',validate='one_to_one')
                 .assign(david_email_dropout = lambda x: x['contacts_drop4'] - x['contacts_drop5'])

)
# seventh dropout: PWC should be excluded
pwc_excl_drop = (pd.concat([tax_nps,legal_nps])
                 .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(ACCNT_EXCL)]
                 .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(accnt_engagement)]
                 .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(ACCNT_INTERNAL)]
                 .loc[lambda x: ~x['EMAIL'].str.contains('|'.join(EMAIL_SENSITIVE))]
                 .loc[lambda x: ~(x['EMAIL'].str.contains('|'.join(EMAIL_INTERNAL)) & x['USERNAME'].str.contains('@'))]
                 .loc[lambda x: ~x['EMAIL'].isin(EMAIL_EXCL)]
                 .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(PWC_EXCL)]
                 .groupby('PRODUCTEMAIL')['USERNAME'].nunique()
                 .reset_index()
                 .rename(columns={'PRODUCTEMAIL':'product','USERNAME':'contacts_drop6'})
                  )

pwc_excl_dropout = (email_excl_dropout
                    .merge(pwc_excl_drop,how='left',on='product',validate='one_to_one')
                    .assign(pwc_dropout = lambda x: x['contacts_drop5'] - x['contacts_drop6'])

)

# split TolleyGuidance_and_TolleyLibrary to two rows and get total Tolley Guidance contacts and Tolley Library contacts respectively
pwc_excl_dropout['product'] = internal_dropout['product'].apply(lambda x: x.split('_and_'))
funnel_view = (pwc_excl_dropout
               .explode('product')
               .groupby('product').sum().reset_index()
               .rename(columns={'contacts_step1':'No_users_usg','contacts_drop1':'No_users_hub_all','contacts_step2':'No_users_hub_email','contacts_drop2':'No_users_excl_acct','contacts_drop3':'No_users_engagement','contacts_drop4':'No_users_MoJ_internal','contacts_drop5':'No_users_email_exclusion','contacts_drop6':'No_users_PWC_exclusion'})             
              )


# final nps file
final_dataset = (pd.concat([tax_nps,legal_nps])
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(ACCNT_EXCL)]
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(accnt_engagement)]
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(ACCNT_INTERNAL)]
                   .loc[lambda x: ~x['EMAIL'].str.contains('|'.join(EMAIL_SENSITIVE))]
                   .loc[lambda x: ~(x['EMAIL'].str.contains('|'.join(EMAIL_INTERNAL)) & x['USERNAME'].str.contains('@'))]
                   .loc[lambda x: ~x['EMAIL'].isin(EMAIL_EXCL)]
                   .loc[lambda x: ~x['GENESIS_ACCOUNT'].isin(PWC_EXCL)]
                   .rename(columns={'PRODUCTEMAIL':'product'})
                  )

final_dataset['product'] = final_dataset['product'].apply(lambda x: x.split('_and_'))
final_dataset = final_dataset.explode('product')
print('generating final dataset ...')
output_final = sys.path[0]+'\\3. Raw Data\\final_data_set.xlsx'
final_dataset.to_excel(output_final,index=False)

# Summary
## number of contacts receiving surveys
no_users_receiving_surveys = (final_dataset.groupby('product')['USERNAME'].count().reset_index()
                              .rename(columns = {'USERNAME':'number_of_contacts'} )
                             )

# Survey Penetration - by segments, how many contacts will receive 1, 2 or 3 surveys
prod_survey = final_dataset['product'].drop_duplicates().tolist()
survey_penetration = (final_dataset.groupby(['VALUE','USERNAME','product'])['CONTACT_ID'].count().unstack().fillna(0).reset_index()
                      .assign(total_surveys = lambda x: x[prod_survey].sum(axis=1))
                      .groupby(['total_surveys','VALUE'])['USERNAME'].count().unstack().sort_values('total_surveys',ascending=False).fillna(0)
                      .reset_index()
                      .rename(columns = {'total_surveys':'number_of_surveys'} )
                      .astype(int)
                       )
survey_penetration.columns = [survey_penetration.columns[0]] + [ 'number_of_contacts_in_' + x for x in survey_penetration.columns if x != 'number_of_surveys']



# save funnel view and summary
print('generating funnel view and summary file ...')
funnel_file = sys.path[0]+'\\3. Raw Data\\funnel_view_and_summary.xlsx'
sheets = ['Funnel_View','No_Contacts','Survey_Penetration']
dfs = [funnel_view,no_users_receiving_surveys,survey_penetration]
wb = openpyxl.load_workbook(funnel_file)

for df, sheet in zip(dfs, sheets):
    if sheet in wb.sheetnames:
        del wb[sheet]
        wb.save(funnel_file)
    writer = pd.ExcelWriter(funnel_file, engine='openpyxl')
    writer.book = wb
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.save()
if 'Sheet1' in wb.sheetnames:
    del wb['Sheet1']
    wb.save(funnel_file)

