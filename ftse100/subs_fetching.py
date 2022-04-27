'''This script is for fetching the below data:
1. all subscriptions to lexis library, lexis psl, and lexis draft & create. note that product family core and legal library services are both for lexis library
2. union output of core and legal library services together for lexis library sub. legal library services is for menus like academic jordan
3. a list of all the accounts ready for mapping
4. active users for all client and law firm

This script also prepares "previous" sheet ready for comparison
'''
import cx_Oracle
import pandas as pd
import csv
import datetime as dt
from datetime import date
#import win32com.client
from pandas import ExcelWriter
import os
import sys
from sqlalchemy.engine import create_engine
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Border
import easygui
import win32ui
import win32con

masterworking = sys.path[0] + '\\masterworking.xlsx'


#--------prepare working sheets----------------------------------------------
print("Prepare the working sheets...")
print("copy previous top200lawfirms data to previous tab...")
wb = openpyxl.load_workbook(masterworking)
sheets = ['Library', 'PSL', 'Draft','Legal Library Services']
previous = wb['previous']
current = wb['Top200LawFirms']
prows = previous[2:previous.max_row]
for prow in prows:
    for cell in prow:
        cell.value = None  # might not need this
        rownumber = cell.row
        columnnumber = cell.column
        currentvalue = current.cell(rownumber, columnnumber).value
        cell.value = currentvalue
wb.save(masterworking)

writer = pd.ExcelWriter(masterworking, engine='openpyxl')
writer.book = wb

# # create an empty sheet for legal library sercives
# df_legal_library_sercives = pd.DataFrame()
# df_legal_library_sercives.to_excel(writer, sheet_name='Legal Library Services', index=False)
# writer.save()

# #------- prepare current master file, delete related sheets for new dataframe------
# print("Delete old raw data sheets...")

# for sheet in sheets:
#     del wb[sheet]
#     wb.save(masterworking)

DIALECT = 'oracle'
SQL_DRIVER = 'cx_oracle'
USERNAME = 'DATAANALYTICS' 
PASSWORD = 'DatPwd123Z' 
HOST = 'PSDB3684.LEXIS-NEXIS.COM'
PORT = 1521
SERVICE = 'GBIPRD1.ISPPROD.LEXISNEXIS.COM'
ENGINE_PATH_WIN_AUTH = DIALECT + '+' + SQL_DRIVER + '://' + USERNAME + ':' + PASSWORD +'@' + HOST + ':' + str(PORT) + '/?service_name=' + SERVICE

engine = create_engine(ENGINE_PATH_WIN_AUTH)
conn_base=engine.connect()

acc_nums = []
acc_names = []
segment = []
cols = []


# Legal Library Services for Jordan menus
prodfam = ['Core', 'PSL/LPA', 'Lexis Check & Draft','Legal Library Services']

df_subs = pd.DataFrame()

#start_date = input("Please type in your start date (<=)'DD_MON-YYYY': ")
start_date = easygui.enterbox("Please type in your start date ( <= )'DD-MON-YYYY', e.g.'31-MAR-2022'")
#effend_date = input("Please type in your effective end date (>=) 'DD-MON-YYYY': ")
effend_date = easygui.enterbox("Please type in your effective end date (>=) 'DD-MON-YYYY', e.g. '01-JAN-2022'")
#time = input("Clarify your date'yyyy-mm-dd: ")
for pf, sheet in zip(prodfam, sheets):
    print("Dumpping data...")
    print(pf)
    query = r'''-- Active subs for each product family within a time period
     with sub as (
 SELECT DISTINCT
        acc.accnt_legcy_id      acc_num,
        acc.accnt_name          acc_name,
        cus.integration_id      cus_id,
        cus.accnt_name          cus_name,
        cus.cust_sub_class_2    market_seg,
        case when PROD.PROD_FAMLY = 'Legal Library Services' then 'Core'
        else PROD.PROD_FAMLY end pf_name,  
        DSUB.SUB_STAT,
        DSUB.DT_START startdate,
        DSUB.DT_END,
        DSUB.CANC_DT,
        CASE
        -- in the past, some cae records were having cae date populated in the canc_dt
    WHEN (DSUB.CANC_DT < DSUB.DT_END and lower(dsub.sub_stat) <>'cancel at end' and extract (year from dsub.canc_dt) <>'1753' ) THEN  DSUB.CANC_DT
    ELSE DSUB.DT_END
    END AS effend
    FROM
        law.f_subscr_rev fsub
        JOIN law.d_subscr_rev     dsub ON fsub.row_wid = dsub.row_wid
        JOIN law.d_fin_accnt_x    acc ON fsub.fa_wid = acc.row_wid
        JOIN law.d_product        prod ON fsub.prod_wid = prod.row_wid
        JOIN law.d_customer_x     cus ON fsub.cus_wid = cus.row_wid
        join law.d_position pos on pos.row_wid = acc.pos_pr_wid
         
    WHERE   1 = 1
        AND fsub.bu_pguid = 'UK'
      
        AND DSUB.TRIAL_FLG <> 'Y'
        --AND dsub.sub_stat not IN ('Pending','Frozen')
        AND  cus.cust_sub_class_2 not in ('Internal','External Test')
        AND lower(dsub.grts_rsn) NOT LIKE '%test%'
        and lower(dsub.grts_rsn) not like '%trial%'

        and pos.POSTN_REP_CD <> '--'
        AND ( acc.accnt_name NOT LIKE '%Internal%'
              AND acc.accnt_name NOT LIKE '%Gratis%' )
              AND acc.accnt_legcy_id NOT IN (
            'HCLT5003',
            'PSLT5000',
            'LEXI5418',
            'HCLT5002',
            'GLOB5164',
            'HCL.5000',
            'WAKI5000',
            'KNOW5038',
            'TREV5076',
            'MSAB5019',
            'LEXI5107'--the real internal account numbers found in Nov 2019
        ) 
        -- Legal Library Services for Jordan menus
        AND PROD.PROD_FAMLY = '{0}' 
        )
    select 
    sub.acc_num,
    sub.acc_name,
    sub.pf_name,
    min(sub.startdate) as minstart,
    max(sub.effend) as maxend,
    sub.market_seg
    from sub
    where 1=1
    and (sub.effend >= '{1}')
    and sub.startdate <= '{2}'
    group by sub.acc_name,sub.pf_name,sub.market_seg,sub.acc_num
    order by 1''' .format(pf,effend_date,start_date)

    df_sub = pd.read_sql(query, con=conn_base)
    cols = df_sub.columns
    # acc_num = df_sub['acc_num'].to_list()
    # acc_name = df_sub['acc_name'].to_list()
    # seg = df_sub['market_seg'].to_list()
    # acc_nums.extend(acc_num)
    # acc_names.extend(acc_name)
    # segment.extend(seg)
    if pf == 'Core':
        df_core = df_sub
    elif pf == 'Legal Library Services':
        df_jordan  = df_sub
    else:
        df_subs = df_subs.append(df_sub)
       # df_sub.to_excel(writer, sheet_name=sheet, index=False)
       # writer.save()

# combine Core and Legal Library Services as they are both LexisLibrary
df_ll = (pd.concat([df_core,df_jordan])
.assign(min_start = lambda x: x.groupby(['acc_num','acc_name','market_seg'])['minstart'].transform('min')) 
.assign(max_end = lambda x: x.groupby(['acc_num','acc_name','market_seg'])['maxend'].transform('max'))
).drop(columns=['minstart','maxend']).rename(columns = {'min_start':'minstart','max_end':'maxend'} )
df_subs = df_subs.append(df_ll[cols])
subscription = sys.path[0] + '\\subscription.xlsx'
df_subs.to_excel(subscription,index = False)
#df_ll[cols].to_excel(writer, sheet_name='Library', index=False)
#writer.save()

# # get a list of de-duped accounts with account number, ready for matching
# accounts = {'account_number': acc_nums,'account_name': acc_names,'segment':segment}
# df_accounts = pd.DataFrame(accounts).drop_duplicates()
# df_accounts.to_excel('./accounts.xlsx',index=False)

#--connect oracle 12c for active users---------------------------------------------------------
startdate = easygui.enterbox("Please put your start date: YYYY / mm (e.g. 2022 / 01): ")
enddate = easygui.enterbox("Please put your end date: YYYY / mm (e.g. 2022 / 03): ")
print("Running query for active users")
query = """with all_client as(select 
  prodfam.pf_name           as prodfam,
  count(distinct   
  (case when prodfam.pf_name  = 'Lexis Check & Draft' and  dusg.usgsrc_type != 'Unspecified' and WU_LOGIN_ID != 'Unspecified'
  then wu.WU_LOGIN_ID
  when prodfam.pf_name  = 'Lexis Check & Draft' and  dusg.usgsrc_type != 'Unspecified' and WU_LOGIN_ID = 'Unspecified'
  and fusg.SRCUSG_VISITOR_REF != 'Unspecified'   then fusg.SRCUSG_VISITOR_REF
  when prodfam.pf_name  != 'Lexis Check & Draft' and wu.wu_login_id != 'Unspecified' then wu.WU_LOGIN_ID end
  )
  ) All_Clients
  
from  LAW.F_SRC_USG_MN_A fusg /* Fact_SourceUsage */
inner join law.d_prod_family prodfam /* Dim_ProductFamily */
on fusg.pf_wid = prodfam.row_wid
inner join law.d_date period /* Dim_Date */
on fusg.dt_wid = period.row_wid
inner join law.d_usg_type usgty  /* Dim_UsageType */
on fusg.usgtp_wid = usgty.row_wid
inner join LAW.D_WEB_USER wu
on fusg.WU_WID = wu.row_wid
inner join law.d_usg_src dusg   /* Dim_UsageSource */
on fusg.usgsrc_wid = dusg.row_wid
where 1  =1
and usgty.usgtp_subtype not in ('COMMIT ADJ', 'PRINT LINES', 'WEB')
and usgty.usgtp_type not    in ('ADF', 'BROWSE', 'Doc Views', 'IMAGE', 'Logins', 'Page Views', 'Visits','Unspecified')
AND prodfam.pf_name IN ('LexisLibrary', 'LexisPSL','Lexis Check & Draft')
and period.yr_mth BETWEEN '{0}' AND '{1}'
group by prodfam.pf_name),

law_firm as(

select 
  prodfam.pf_name           as prodfam,
   count(distinct    
  (case when prodfam.pf_name  = 'Lexis Check & Draft' and  dusg.usgsrc_type != 'Unspecified' and WU_LOGIN_ID != 'Unspecified'
  then wu.WU_LOGIN_ID
  when prodfam.pf_name  = 'Lexis Check & Draft' and  dusg.usgsrc_type != 'Unspecified' and WU_LOGIN_ID = 'Unspecified'
  and fusg.SRCUSG_VISITOR_REF != 'Unspecified'   then fusg.SRCUSG_VISITOR_REF
  when prodfam.pf_name  != 'Lexis Check & Draft' and wu.wu_login_id != 'Unspecified' then wu.WU_LOGIN_ID end
  )
  )  Law_Firms  
from  LAW.F_SRC_USG_MN_A fusg /* Fact_SourceUsage */
inner join law.d_prod_family prodfam /* Dim_ProductFamily */
on fusg.pf_wid = prodfam.row_wid
inner join law.d_date period /* Dim_Date */
on fusg.dt_wid = period.row_wid
inner join law.d_customer_x seg  /* Dim_Customer_Legacy */
on fusg.cus_wid = seg.row_wid
inner join law.d_usg_type usgty  /* Dim_UsageType */
on fusg.usgtp_wid = usgty.row_wid
inner join LAW.D_WEB_USER wu
on fusg.WU_WID = wu.row_wid
inner join law.d_usg_src dusg   /* Dim_UsageSource */
on fusg.usgsrc_wid = dusg.row_wid
where 1  =1
and usgty.usgtp_subtype not in ('COMMIT ADJ', 'PRINT LINES', 'WEB')
and usgty.usgtp_type not    in ('ADF', 'BROWSE', 'Doc Views', 'IMAGE', 'Logins', 'Page Views', 'Visits','Unspecified')
AND prodfam.pf_name IN ('LexisLibrary', 'LexisPSL','Lexis Check & Draft')
and seg.CUST_SUB_CLASS_2 in ('Top Tier','Full Service Commercial','General Practice','Consumer-led','Small Law')
and period.yr_mth BETWEEN '{0}' AND '{1}'
group by prodfam.pf_name)

select 
ac.prodfam,
ac.all_clients,
lf.Law_Firms
from all_client ac
join law_firm lf
on ac.prodfam = lf.prodfam

""".format(startdate, enddate)

df_active = pd.read_sql(query, con=conn_base)

#--dump the data to masterworking file-------------------------------------
print("dumping active user data...")
del wb['activeusers']
df_active.to_excel(writer, sheet_name="activeusers", index=False)
writer.save()
wb.save(masterworking)

conn_base.close()

print("Data loaded")



