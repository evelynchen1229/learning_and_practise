import cx_Oracle
import pandas as pd
import csv
import datetime as dt
from datetime import date
#import win32com.client
from pandas import ExcelWriter
import os
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Border
import easygui
import win32ui
import win32con
import numpy as np
import datetime as dt
from datetime import date

masterworking = sys.path[0] + '\\masterworking.xlsx'
subscription = sys.path[0] + '\\subscription.xlsx'
mapping = sys.path[0] + '\\mapping.xlsx'
law_list=sys.path[0] + '\\top200law.xlsx'

df_mapping_law = pd.read_excel(mapping,sheet_name='Lawyer200')
df_mapping_ftse = pd.read_excel(mapping,sheet_name='FTSE100')
df_subs = pd.read_excel(subscription)

magic_circle = ['Clifford Chance','Allen & Overy','Linklaters','Freshfields Bruckhaus Deringer','Slaughter and May']

def quarter(mth):
    if mth == 0:
        return 'Q4'
    else:
        quarter_num = mth // 4 + 1
        return 'Q' + str(quarter_num)

def year(yr,mth):
    if mth == 0:
        return str(yr - 1)
    else:
        return str(yr)

def prev_yr_quarter(curr_yr,curr_qt):
    if curr_qt == 'Q1':
        prev_yr = str(int(curr_yr) - 1)
        prev_qt = 'Q4'
    else:
        prev_yr = curr_yr
        prev_qt = 'Q' + str(int(curr_qt[-1]) - 1)
    return prev_yr + ' ' + prev_qt

def toplawcus(x,col = 'uk 200'):
    cus_list = x[col][x[col].fillna(0) != 0].unique()
    return cus_list

def counting(df,pf,column,mc=False):
    if mc == False:
        number_of_top200_ftse_list = df[df[pf]=='Y'].groupby(['Class'])[column].count().values
    else:
        number_of_top200_ftse_list = df[(df[pf]=='Y') & df['Firm'].isin(magic_circle)].groupby(['Class'])[column].count().values
    return number_of_top200_ftse_list


today_yr = date.today().year
today_mth = date.today().month
reporting_quarter = quarter(today_mth - 1)
reporting_year = year(today_yr, today_mth - 1)
ftse_list = sys.path[0] + f'\\{reporting_year}\\{reporting_year} {reporting_quarter}\\ftse100_{reporting_year}{reporting_quarter}.csv'


#--------prepare working sheets----------------------------------------------

wb = openpyxl.load_workbook(masterworking)
writer = pd.ExcelWriter(masterworking, engine='openpyxl')
writer.book = wb

df_law_for_reporting = (df_mapping_law[['ACC_NUM','UK 200 Rank','UK 200']]
.loc[lambda x: x['UK 200 Rank'].fillna(0) != 0]
)
df_law_for_reporting.columns = df_law_for_reporting.columns.str.lower()

df_ftse_for_reporting = (df_mapping_ftse.loc[lambda x: x['Comment'].fillna(0) == 0]
)[['ACC_NUM','FTSE100','FTSE100 Code']]
df_ftse_for_reporting.columns = df_ftse_for_reporting.columns.str.lower()

df_for_reporting = (df_subs.merge(df_law_for_reporting,how='left',on='acc_num')
.merge(df_ftse_for_reporting,how = 'left',on='acc_num')
)

df_library = df_for_reporting.loc[lambda x: x['pf_name'] == 'Core']
df_psl = df_for_reporting.loc[lambda x: x['pf_name'] == 'PSL/LPA']
df_draft = df_for_reporting.loc[lambda x: x['pf_name'] == 'Lexis Check & Draft']


sheets = ['Library', 'PSL', 'Draft']
dfs = [df_library,df_psl,df_draft]

for df, sheet in zip(dfs, sheets):
    del wb[sheet]
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.save()

wb.save(masterworking)


# -- put in ftse and toplaww lists---

df_law = (pd.read_excel(law_list)[['Rank','Firm']].loc[lambda x: x['Rank'].fillna(0) != 0].drop_duplicates()
.assign(Class = lambda x: np.where(x['Rank'] <= 100, 'Top 100 Law Firms','Top 200 Law Firms'))
.assign(Library = lambda x: np.where(x['Firm'].isin(toplawcus(df_library)),'Y','N'))
.assign(PSL = lambda x: np.where(x['Firm'].isin(toplawcus(df_psl)),'Y','N'))
.assign(Draft = lambda x: np.where(x['Firm'].isin(toplawcus(df_draft)),'Y','N'))
)

df_ftse = (pd.read_csv(ftse_list)[['Code','Original Name']]
.assign(Class = 'FTSE 100')
.assign(Library = lambda x: np.where(x['Code'].isin(toplawcus(df_library,col='ftse100 code')),'Y','N'))
.assign(PSL = lambda x: np.where(x['Code'].isin(toplawcus(df_psl,col='ftse100 code')),'Y','N'))
.assign(Draft = lambda x: np.where(x['Code'].isin(toplawcus(df_draft,col='ftse100 code')),'Y','N'))
.rename(columns = {'Original Name':'Name'})
)

df_lists = [df_law,df_ftse]
list_sheets = ['Breakdown - Top200LawFirms','Breakdown - Top100FTSE']
for df_list,list_sheet in zip(df_lists,list_sheets):
    del wb[list_sheet]
    df_list.to_excel(writer,sheet_name = list_sheet,index=False)
    writer.save()
top100 = []
top200 = []
law_magic_circle = []
ftse100 = []

for sheet in sheets:
    toplaw = counting(df_law,sheet,'Firm')
    top100.append(toplaw[0])
    top200.append(toplaw[0] + toplaw[1])
    toplaw_magic_circle = counting(df_law,sheet,'Firm',True)
    law_magic_circle.append(toplaw_magic_circle[0])
    ftse = counting(df_ftse,sheet,'Code')
    ftse100.append(ftse[0])

# ----- writing in the summary page and do the actual flag-----

Summary = wb['Summary']
Summary_range_list = [Summary["C5":"E8"], Summary["C12":"E16"], Summary["C21":"D23"]]
for Summary_range in Summary_range_list:
    for Summary_cell in Summary_range:
        for s in Summary_cell:
            s.value = None
column_start = 3
column_end = 5
for column_number in range(column_start, column_end + 1):
    Summary.cell(5,column_number).value = top100[column_number-3]
    Summary.cell(6,column_number).value = top200[column_number-3]
    Summary.cell(7,column_number).value = law_magic_circle[column_number-3]
    Summary.cell(8,column_number).value = ftse100[column_number-3]

n = 0
row_num = 12
for sheet,column_number in zip(sheets,range(column_start, column_end + 1)):
    while n < 5 : 
        Summary.cell(row_num + n,column_number).value  = df_law[df_law['Firm'] == magic_circle[n]][sheet].values[0]
        n += 1
    n = 0

activeuser = wb['activeusers']
Summary.cell(21,3).value=activeuser.cell(2,2).value
Summary.cell(22, 3).value = activeuser.cell(3, 2).value
Summary.cell(23,3).value=activeuser.cell(4,2).value
Summary.cell(21,4).value=activeuser.cell(2,3).value
Summary.cell(22,4).value=activeuser.cell(3,3).value
Summary.cell(23, 4).value = activeuser.cell(4, 3).value

wb.save(masterworking)