import pandas as pd
import requests 
import urllib.request
import urllib
import os
import sys
import datetime as dt
from datetime import date
from bs4 import BeautifulSoup
import csv
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Border
import numpy as np

def quarter(mth):
    if mth == 0:
        return 'Q4'
    else:
        quarter_num = mth // 4 + 1
        return 'Q' + str(quarter_num)

def year(yr,mth):
    if mth == 0:
        return str(yr - 1)
    else:
        return str(yr)

def prev_yr_quarter(curr_yr,curr_qt):
    if curr_qt == 'Q1':
        prev_yr = str(int(curr_yr) - 1)
        prev_qt = 'Q4'
    else:
        prev_yr = curr_yr
        prev_qt = 'Q' + str(int(curr_qt[-1]) - 1)
    return prev_yr + ' ' + prev_qt

# to get the latest ftse100 list
def ftse():
    pages = [1,2,3,4,5]
    pages += pages [::-1]
    codes = []
    names = []
    original_names = []
    # FTSE100 website is designed in a way that the information is repetitive when you go to each page (seems to be fixed 20 companies per page)
    # In order not to miss out any company, this script will go all over the pages again
    for page in pages:
        url = f'https://www.londonstockexchange.com/indices/ftse-100/constituents/table?page={page}'
        response = requests.get(url)
        soup = BeautifulSoup(response.text,'html.parser')
        table = soup.find('table',{'class':'full-width ftse-index-table-table'})
        table_body = table.findAll('tbody')[0]
        table_rows = table_body.findAll('tr')
        for tr in table_rows:
            href = tr.find('a').get('href')
            href = href.split('/')
            code = href[1]
            original_name = href[-1].replace('b-m-european-value-retail-s-a','B&M European Value Retail S.A').replace('smith-ds-plc','SMITH (DS) PLC').replace('international-consolidated-airlines-group-s-a','INTERNATIONAL CONSOLIDATED AIRLINES GROUP S.A.').replace('st-james-s-place-plc',"ST. JAMES'S PLACE PLC").replace('sainsbury-j-plc','SAINSBURY (J) PLC').replace('m-g-plc','M&G PLC').replace('-',' ').replace('coca cola hbc ag','COCA-COLA HBC AG').upper()
            name = href[-1].replace('-',' ').replace('hbc ag','').replace('holdings the plc','').replace('group plc','').replace('group holdings','').replace('plc','').replace('holdings','').replace('ltd','').strip()
            name = name.replace('sainsbury j','sainsbury').replace('international consolidated airlines group s a','international consolidated airlines').replace('legal general','legal and general').replace('st james s place',"st james's place").replace('m g','m and g').replace('smith ds','ds smith').replace('b m european value retail s a','b and m european value retail').replace('coca cola','coca-cola')
            codes.append(code)
            names.append(name)
            original_names.append(original_name)
    ftse100 = pd.DataFrame({'Code':codes,'Name':names,'Original Name':original_names}).drop_duplicates()
    output = f'./{reporting_year}/{reporting_year} {reporting_quarter}/ftse100_{reporting_year}{reporting_quarter}.csv'
    ftse100.to_csv(output, index=False)
    num_codes = len(ftse100['Code'].tolist())
    print('There are 100 FTSE100 companies in the file:',num_codes == 100, f'; Number of FTSE100 companies in the file: {num_codes}')
    # Unless there is Royal Dutch Shell in the FTSE100 list, you should expect just 100 companies. Royal Dutch Shell normally has A and B securities (RDSA and RDSB)
    return num_codes >= 100

today_yr = date.today().year
today_mth = date.today().month
reporting_quarter = quarter(today_mth - 1)
reporting_year = year(today_yr, today_mth - 1)

num_ftse_codes_as_expected = False
# Due to the way London exchange website is designed, you might need to keep running this script for a few times to get all the companies
# If running too long, can use ctrl + c or ctrl + z to interrupt
while num_ftse_codes_as_expected == False:
    num_ftse_codes_as_expected = ftse()



'''running comparison against the list for last quarter'''

prev_reporting_year_quarter = prev_yr_quarter(reporting_year,reporting_quarter)

'''load dataframes'''

curr_ftse = sys.path[0] +f'\\{reporting_year}\\{reporting_year} {reporting_quarter}\\ftse100_{reporting_year}{reporting_quarter}.csv'
prev_ftse = sys.path[0] +f'\\{prev_reporting_year_quarter[:4]}\\{prev_reporting_year_quarter}\\ftse100_{prev_reporting_year_quarter.replace(" ","")}.csv'

df_curr_ftse = pd.read_csv(curr_ftse)
df_prev_ftse = pd.read_csv(prev_ftse)

'''check whether any changes in codes / names'''
'''1. generating dataframes'''
df_ftse_drop_out = (df_prev_ftse.merge(df_curr_ftse,how='left',on='Code',suffixes=('_prev','_curr'))
.loc[lambda x: pd.isna(x['Name_curr'])]
)
df_ftse_new_code = (df_curr_ftse.merge(df_prev_ftse,how='left',on='Code',suffixes=('_curr','_prev'))
.loc[lambda x: pd.isna(x['Name_prev'])]
)
# this is just for same code different names - for code and name change, it needs to be manually check and added to the sheet (apart from shell)
df_ftse_name_change = (df_curr_ftse.merge(df_prev_ftse,how='inner',on='Code',suffixes=('_curr','_prev'))
.loc[lambda x: x['Name_curr'] != x['Name_prev']]
.rename(columns = {'Code':'Code_prev'})
.assign(Code_curr = np.nan)
)[['Code_curr','Code_prev','Name_curr','Name_prev']]

if 'shell' in df_ftse_new_code['Name_curr'].to_list():
    new_code = ['SHEL','SHEL']
    old_code = ['RDSA','RDSB']
    new_name = ['shell','shell']
    old_name = ['royal dutch shell','royal dutch shell']
    change_dict = {'Code_curr':new_code,'Code_prev':old_code,'Name_curr':new_name,'Name_prev':old_name}
    df_change = pd.DataFrame(data=change_dict)
    df_ftse_name_change = df_ftse_name_change.append(df_change)



'''2. producing file'''
print('producing file')
comparison = sys.path[0] + f'\\{reporting_year}\\{reporting_year} {reporting_quarter}\\ftse100_comparison.xlsx'
dfs = [df_ftse_drop_out,df_ftse_new_code,df_ftse_name_change]
sheets = ['ftse_drop_out', 'ftse_new_code', 'df_ftse_name_change']
try:
    comparison_file = pd.read_excel(comparison)
except FileNotFoundError:
    wb = Workbook()
    for sheet in sheets:
        ws = wb.create_sheet(sheet)
    del wb['Sheet']
    wb.save(filename = comparison)
wb = openpyxl.load_workbook(comparison)

writer = pd.ExcelWriter(comparison, engine='openpyxl')
writer.book = wb

for df, sheet in zip(dfs, sheets):
    del wb[sheet]
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.save()


wb.save(comparison)

